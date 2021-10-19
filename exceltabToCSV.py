# Kate Bowers (adapted from billydh on Github)
# Levin Lab Tufts University
# Program development started March 2021
# Beckman Scholars - Tadpole Bullying
# exceltabToCSV.py -- Extracts each sheet from an Excel workbook as a CSV file

import csv
import xlrd
import openpyxl as op
import os

# def excelToCSV(excel_file, csv_file_base_path):  # adapted from exceltab_to_csv by Github user billydh
#     print("excel split starting now!")
#     print("exc file is " + excel_file)
#     print("path is " + csv_file_base_path)
#     # print(os.path.isfile(excel_file))
#     # workbook = xlrd.open_workbook(excel_file)
#     print("trying to open book lol ")
#     workbook = op.load_workbook(filename = excel_file, read_only=True)
#     print("opened workbook!")
#     print(workbook.sheetnames)
#     print("yep")
#     for sheet_name in workbook.sheetnames:
#         print('processing - ' + sheet_name)
#         worksheet = workbook[sheet_name]
#         csv_file_full_path = csv_file_base_path + sheet_name.lower().replace(" - ", "_").replace(" ", "_") + '.csv'
#         csvfile = open(csv_file_full_path, 'w')
#
#         writetocsv = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
#         for rownum in range(worksheet.max_row):
#             writetocsv.writerow([x for x in worksheet[row_num]])  # does this work>
#             print("grabbed row ", rownum)
#         csvfile.close()
#         print(sheet_name + ' has been saved at - ' + csv_file_full_path)
#     workbook.close()

def excelToCSV(excel_file, csv_file_base_path):  # adapted from exceltab_to_csv by Github user billydh
    print("starting now")
    print("exc file is " + excel_file)
    print("path is " + csv_file_base_path)

    workbook = op.load_workbook(filename=excel_file, read_only=True)
    print("opened workbook!")
    print(workbook.sheetnames)
    print("printed sheetnames")
    for sheet_name in workbook.sheetnames:
        print('processing - ' + sheet_name)
        worksheet = workbook[sheet_name]
        csv_file_full_path = csv_file_base_path + sheet_name.lower().replace(" - ", "_").replace(" ", "_") + '.csv'
        csvfile = open(csv_file_full_path, 'w')

        c = csv.writer(csvfile, quoting=csv.QUOTE_ALL)
        for row in worksheet.rows:
            c.writerow([cell.value for cell in row])  # does this work>
        csvfile.close()
        print(sheet_name + ' has been saved at - ' + csv_file_full_path)
    workbook.close()